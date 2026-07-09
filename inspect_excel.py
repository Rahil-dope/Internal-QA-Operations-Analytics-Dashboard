import openpyxl
import datetime

wb = openpyxl.load_workbook('final review file.xlsx', data_only=True)
for sheet in wb.sheetnames:
    ws = wb[sheet]
    print("=" * 60)
    print(f"Sheet: {sheet}")
    print(f"Dimensions: {ws.dimensions}")
    
    # Read rows
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        print("Empty sheet")
        continue
        
    print(f"Total Rows: {len(rows)}")
    
    # We find headers. Sometimes there might be empty rows at the top (like Shrinkage which had None values in row 1)
    # Let's inspect first 5 rows to see where the headers actually start.
    print("First 5 rows:")
    for idx, r in enumerate(rows[:5]):
        non_none = [v for v in r if v is not None]
        print(f"  Row {idx+1} ({len(non_none)} non-None): {r[:10]}...")

    # Let's print out what we detect as headers and types of the first data row
    # For Shrinkage, Row 1 had mostly None except Hyderabad/Overall, Row 2 had headers.
    # Let's see how we can generalise this.
