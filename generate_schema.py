import openpyxl
import datetime

wb = openpyxl.load_workbook('final review file.xlsx', data_only=True)
with open('excel_schema.txt', 'w', encoding='utf-8') as f:
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        f.write("=" * 60 + "\n")
        f.write(f"Sheet: {sheet}\n")
        f.write(f"Dimensions: {ws.dimensions}\n")
        
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            f.write("Empty sheet\n")
            continue
            
        f.write(f"Total Rows: {len(rows)}\n")
        
        # Determine header index (for Shrinkage, first non-empty header row is row 2)
        header_row_idx = 0
        if sheet == 'Shrinkage':
            header_row_idx = 1
            
        headers = rows[header_row_idx]
        data_rows = rows[header_row_idx + 1:]
        
        f.write(f"Headers (index-based):\n")
        for col_idx, h in enumerate(headers):
            sample_val = None
            if data_rows:
                sample_val = data_rows[0][col_idx]
            val_type = type(sample_val).__name__ if sample_val is not None else "None"
            f.write(f"  Col {col_idx}: {h} (Sample Type: {val_type}, Sample Value: {sample_val})\n")
        f.write("\n")

print("Schema generated in excel_schema.txt")
